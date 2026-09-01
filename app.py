import random
import io

from flask import Flask, render_template, session, send_file, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


app = Flask(__name__)
app.secret_key = "losowanie_grup_2026"


lista = [
    "Adam i Wanda",
    "Stas_K",
    "Marta i Antoni_K..",
    "Marcin",
    "Michał i Zuzanna",
    "Ola_P.",
    "Wojciech i Judyta Sch.",
    "Łukasz i Patrycja",
    "Stasiu i Ania",
    "Franek i Judyta Sz.",
    "Gabriela",
    "Grzesiu",
    "Krzychu",
    "Jacek i Ida",
    "Teresa",
    "Marek i Małgorzata",
    "Piotr",
    "Maria_P",
    "Marta_P",
    "Magda"
]


def losuj_grupy():

    single = []
    pary = []

    for osoba in lista:
        if len(osoba) <= 8:
            single.append(osoba)
        else:
            pary.append(osoba)

    random.shuffle(single)
    random.shuffle(pary)

    grupy = [[], [], [], []]

    # Rozdzielamy małżeństwa
    for i, para in enumerate(pary):
        numer_grupy = i % 4
        grupy[numer_grupy].append(para)

    # Rozdzielamy singli
    for osoba in single:

        najmniejsza_grupa = min(
            grupy,
            key=lambda grupa: sum(
                2 if element in pary else 1
                for element in grupa
            )
        )

        najmniejsza_grupa.append(osoba)

    return grupy


@app.route("/")
def strona_glowna():

    return render_template(
        "index.html",
        grupy=None
    )


@app.route("/losuj")
def losuj():

    grupy = losuj_grupy()

    session["grupy"] = grupy

    return render_template(
        "index.html",
        grupy=grupy
    )


@app.route("/pobierz")
def pobierz():

    grupy = session.get("grupy")

    if not grupy:
        return redirect(url_for("strona_glowna"))

    # =============================
    # TWORZENIE PLIKU EXCEL
    # =============================

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Wylosowane grupy"

    # Kolory nagłówków
    kolory = [
        "4F81BD",
        "9BBB59",
        "F79646",
        "8064A2"
    ]

    cienka_linia = Side(
        style="thin",
        color="CCCCCC"
    )

    # =============================
    # NAGŁÓWKI GRUP
    # =============================

    for numer_grupy in range(4):

        kolumna = numer_grupy + 1

        komorka = sheet.cell(
            row=1,
            column=kolumna
        )

        komorka.value = f"GRUPA {numer_grupy + 1}"

        komorka.font = Font(
            bold=True,
            color="FFFFFF",
            size=14
        )

        komorka.fill = PatternFill(
            start_color=kolory[numer_grupy],
            end_color=kolory[numer_grupy],
            fill_type="solid"
        )

        komorka.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        komorka.border = Border(
            left=cienka_linia,
            right=cienka_linia,
            top=cienka_linia,
            bottom=cienka_linia
        )

    # =============================
    # OSOBY W GRUPACH
    # =============================

    for numer_grupy, grupa in enumerate(grupy):

        kolumna = numer_grupy + 1

        for numer_wiersza, osoba in enumerate(
            grupa,
            start=2
        ):

            komorka = sheet.cell(
                row=numer_wiersza,
                column=kolumna
            )

            komorka.value = osoba

            komorka.font = Font(
                size=12
            )

            komorka.alignment = Alignment(
                horizontal="left",
                vertical="center"
            )

            komorka.border = Border(
                left=cienka_linia,
                right=cienka_linia,
                top=cienka_linia,
                bottom=cienka_linia
            )

    # =============================
    # SZEROKOŚĆ KOLUMN
    # =============================

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["D"].width = 28

    # Większy nagłówek
    sheet.row_dimensions[1].height = 30

    # =============================
    # ZAPIS DO PAMIĘCI
    # =============================

    plik = io.BytesIO()

    workbook.save(plik)

    plik.seek(0)

    return send_file(
        plik,
        as_attachment=True,
        download_name="wylosowane_grupy.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    app.run(debug=True)